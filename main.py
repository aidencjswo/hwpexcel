from openpyxl import Workbook

wb = Workbook()

#select by worksheet title
ws1 = wb.get_sheet_by_name('Sheet')

#ws1 title rename
ws1.title = 'aiden'

#두번 째 워크시트 생성
ws2 = wb.create_sheet('aiden2')

print(wb.sheetnames)

#save file
wb.save('balances.xlsx')